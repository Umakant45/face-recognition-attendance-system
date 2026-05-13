import cv2
import numpy as np
import os
import sys
import logging
from datetime import datetime
from openpyxl import Workbook, load_workbook

# ── Configuration ─────────────────────────────────────────────────────────────
MODEL_PATH      = "trainer.yml"
DATASET_PATH    = "dataset"
EXCEL_FILE      = "attendance.xlsx"
CONFIDENCE_THRESHOLD = 50   # Lower = stricter match (0–100 scale)
LOG_FILE        = "attendance.log"

# ── Logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)s  %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger(__name__)

# ── Load model ────────────────────────────────────────────────────────────────
if not os.path.exists(MODEL_PATH):
    log.error(f"Model file '{MODEL_PATH}' not found. Run train.py first.")
    sys.exit(1)

recognizer = cv2.face.LBPHFaceRecognizer_create()
recognizer.read(MODEL_PATH)

detector = cv2.CascadeClassifier(
    cv2.data.haarcascades + "haarcascade_frontalface_default.xml"
)

# ── User map ──────────────────────────────────────────────────────────────────
def get_users(path=DATASET_PATH):
    """Build {id: name} from dataset filenames. Each ID mapped once."""
    users = {}
    if not os.path.exists(path):
        return users
    for file in os.listdir(path):
        ext = os.path.splitext(file)[1].lower()
        if ext not in {".jpg", ".jpeg", ".png"}:
            continue
        parts = file.split("_")
        if len(parts) < 2:
            continue
        try:
            uid  = int(parts[0])
            name = parts[1]
            users[uid] = name
        except ValueError:
            pass
    return users

users = get_users()
if not users:
    log.warning("No registered users found in dataset. Attendance will show 'Unknown'.")

# ── Excel setup ───────────────────────────────────────────────────────────────
def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"
        ws.append(["ID", "Name", "Date", "Time"])
        wb.save(EXCEL_FILE)
        log.info(f"Created new attendance file: {EXCEL_FILE}")

init_excel()

# ── Attendance logic ──────────────────────────────────────────────────────────
# In-memory set to avoid re-reading Excel for every face detection
_marked_today: set[int] = set()

def _load_today_marked():
    """Populate _marked_today from existing Excel rows for today's date."""
    today = datetime.now().strftime("%Y-%m-%d")
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[2] == today and row[0] is not None:
            _marked_today.add(int(row[0]))

_load_today_marked()

def mark_attendance(user_id: int, user_name: str):
    if user_id in _marked_today:
        return  # Already marked — no need to open the file

    now = datetime.now()
    date_str = now.strftime("%Y-%m-%d")
    time_str = now.strftime("%H:%M:%S")

    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        ws.append([user_id, user_name, date_str, time_str])
        wb.save(EXCEL_FILE)
        _marked_today.add(user_id)
        log.info(f"Attendance marked — ID:{user_id} Name:{user_name} at {time_str}")
    except Exception as e:
        log.error(f"Failed to write attendance for {user_name}: {e}")

# ── Camera ────────────────────────────────────────────────────────────────────
cam = cv2.VideoCapture(0)
if not cam.isOpened():
    log.error("Cannot open camera. Exiting.")
    sys.exit(1)

cam.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
cam.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)

font       = cv2.FONT_HERSHEY_SIMPLEX
GREEN      = (0, 255, 0)
RED        = (0, 0, 255)
BLUE       = (255, 120, 0)
WHITE      = (255, 255, 255)

log.info("Attendance system started. Press ESC to quit.")
print("\nPress ESC to stop.\n")

while True:
    ret, frame = cam.read()
    if not ret:
        log.error("Failed to grab frame from camera.")
        break

    gray  = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    faces = detector.detectMultiScale(gray, scaleFactor=1.2, minNeighbors=5)

    for (x, y, w, h) in faces:
        face_roi = gray[y:y + h, x:x + w]
        face_id, confidence = recognizer.predict(face_roi)

        if confidence < CONFIDENCE_THRESHOLD:
            name  = users.get(face_id, f"ID-{face_id}")
            label = f"{name}  ({confidence:.0f}%)"
            color = GREEN
            mark_attendance(face_id, name)

            # Tick mark if already marked
            status = " [marked]" if face_id in _marked_today else ""
            label += status
        else:
            label = "Unknown"
            color = RED

        # Bounding box + label
        cv2.rectangle(frame, (x, y), (x + w, y + h), color, 2)
        # Background pill for text
        (tw, th), _ = cv2.getTextSize(label, font, 0.65, 2)
        cv2.rectangle(frame, (x, y - th - 10), (x + tw + 4, y), color, -1)
        cv2.putText(frame, label, (x + 2, y - 5), font, 0.65, WHITE, 2)

    # HUD
    today_str  = datetime.now().strftime("%Y-%m-%d  %H:%M:%S")
    marked_str = f"Marked today: {len(_marked_today)}"
    cv2.putText(frame, today_str,  (10, 25),  font, 0.6, BLUE, 2)
    cv2.putText(frame, marked_str, (10, 50),  font, 0.6, BLUE, 2)

    cv2.imshow("Face Recognition Attendance — ESC to quit", frame)

    if cv2.waitKey(1) & 0xFF == 27:  # ESC
        log.info("Attendance system stopped by user.")
        break

cam.release()
cv2.destroyAllWindows()
